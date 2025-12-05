# -*- coding: utf-8 -*-
"""
生成Excel测试数据文件
运行此脚本将创建包含测试用例的Excel文件
"""
import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from config.config import TEST_DATA_DIR, TEST_DATA_FILE

def create_test_data_excel():
    """创建测试数据Excel文件"""
    
    # 确保目录存在
    os.makedirs(TEST_DATA_DIR, exist_ok=True)
    
    # 创建工作簿
    wb = Workbook()
    ws = wb.active
    ws.title = "登录测试用例"
    
    # 设置表头
    headers = [
        'case_id',      # 用例ID
        'case_name',    # 用例名称
        'username',     # 用户名
        'password',     # 密码
        'expected',     # 预期结果 (success/fail)
        'description',  # 用例描述
        'run'          # 是否执行 (yes/no)
    ]
    
    # 写入表头并设置样式
    header_fill = PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=11)
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # 测试数据
    test_cases = [
        # case_id, case_name, username, password, expected, description, run
        ['TC001', '正确的账号和密码', 'test1', 'test@user.c', 'success', '使用正确的超级管理员账号和密码登录', 'yes'],
        ['TC002', '错误的账号', 'wronguser', 'Ld@513.c', 'fail', '使用错误的账号，正确的密码', 'yes'],
        ['TC003', '错误的密码', 'test1', 'wrongpassword', 'fail', '使用正确的账号，错误的密码', 'yes'],
        ['TC004', '账号和密码都错误', 'wronguser', 'wrongpassword', 'fail', '使用错误的账号和错误的密码', 'yes'],
        ['TC005', '空账号', '', 'user123456', 'fail', '账号为空，密码正确', 'yes'],
        ['TC006', '空密码', 'admin', '', 'fail', '账号正确，密码为空', 'yes'],
        ['TC007', '账号和密码都为空', '', '', 'fail', '账号和密码都为空', 'yes'],
        ['TC008', '账号包含特殊字符', 'admin@123', 'Ld@513.c', 'fail', '账号包含特殊字符', 'yes'],
        ['TC009', '密码包含空格', 'admin', 'user 123456', 'fail', '密码包含空格', 'yes'],
        ['TC010', '账号大小写错误', 'ADMIN', 'user123456', 'fail', '账号大小写错误', 'yes'],
    ]
    
    # 写入测试数据
    for row_num, case_data in enumerate(test_cases, 2):
        for col_num, value in enumerate(case_data, 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = value
            cell.alignment = Alignment(horizontal='left', vertical='center')
    
    # 调整列宽
    column_widths = [12, 25, 15, 15, 12, 35, 10]
    for col_num, width in enumerate(column_widths, 1):
        ws.column_dimensions[chr(64 + col_num)].width = width
    
    # 保存文件
    wb.save(TEST_DATA_FILE)
    print(f"测试数据文件已创建: {TEST_DATA_FILE}")
    print(f"\nExcel表结构说明:")
    print("=" * 80)
    print(f"{'列名':<15} {'说明':<50}")
    print("-" * 80)
    print(f"{'case_id':<15} 用例ID，唯一标识每个测试用例")
    print(f"{'case_name':<15} 用例名称，简短描述测试场景")
    print(f"{'username':<15} 测试用的用户名")
    print(f"{'password':<15} 测试用的密码")
    print(f"{'expected':<15} 预期结果: 'success'表示登录成功, 'fail'表示登录失败")
    print(f"{'description':<15} 用例的详细描述")
    print(f"{'run':<15} 是否执行该用例: 'yes'执行, 'no'跳过")
    print("=" * 80)

if __name__ == '__main__':
    create_test_data_excel()
